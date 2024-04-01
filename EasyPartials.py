#!/usr/bin/python3
# Author: Anthony Garrett
#
# Small script that will read in inventory location data from one spreadsheet
# and transfer that data to another spreadsheet
#
import os
from openpyxl import Workbook, load_workbook

ORIGINAL_INPUT = "EXPORT.XLSX"
DATA_INPUT = "data.xlsx"

def main():
    data_wb = load_workbook(DATA_INPUT)
    data_sheet = data_wb.active

    data_products = data_sheet["A"]
    data_full_quantities = data_sheet["B"]
    
    wb = load_workbook(ORIGINAL_INPUT)
    sheet = wb.active

    # Deleting the first row which contains headers for the columns
    sheet.delete_rows(1, 1)

    # Selecting the columns that hold the needed information about each location
    location_codes = sheet["A"]
    actual_quantity = sheet["D"]
    handling_units = sheet["G"]

    # Parsing out the aisle number from the location code
    code_parts = location_codes[0].value.split("-")[0]

    data_core = create_dictionary(int(code_parts))
    OUTPUT_FILENAME = "Aisle-" + code_parts + "-totals" + ".xlsx"

    # Cleaning excess spaces that are at the beginning of the location codes
    # in the cells
    for code in location_codes:
        code.value = code.value.strip()

    for count, code in enumerate(location_codes):
        data_core[code.value][0] += 1

        # Special case for pallets that are totes holding up to 3000 hot pockets
        # These are systematically entered as having 96 cases
        # No pallet has more than 220 cases except for these special tote
        # pallets
        if actual_quantity[count].value > 220:
            data_core[code.value][1] += 96
        else:
            data_core[code.value][1] += actual_quantity[count].value

    out_book = Workbook()

    out_sheet = out_book.active

    count = [1, 1, 1, 1, 1, 1]

    for key, value in data_core.items():
        find_level = key.split("-")
        level = find_level[2]

        if str(level) == "A":
            out_sheet["A" + str(count[0])] = key
            out_sheet["B" + str(count[0])] = value[1]
            out_sheet["C" + str(count[0])] = value[0]
            count[0] += 1

        elif str(level) == "B":
            out_sheet["E" + str(count[1])] = key
            out_sheet["F" + str(count[1])] = value[1]
            out_sheet["G" + str(count[1])] = value[0]
            count[1] += 1

        elif str(level) == "C":
            out_sheet["I" + str(count[2])] = key
            out_sheet["J" + str(count[2])] = value[1]
            out_sheet["K" + str(count[2])] = value[0]
            count[2] += 1

        elif str(level) == "D":
            out_sheet["M" + str(count[3])] = key
            out_sheet["N" + str(count[3])] = value[1]
            out_sheet["O" + str(count[3])] = value[0]
            count[3] += 1

        elif str(level) == "E":
            out_sheet["Q" + str(count[4])] = key
            out_sheet["R" + str(count[4])] = value[1]
            out_sheet["S" + str(count[4])] = value[0]
            count[4] += 1

            if int(find_level[1]) == 18:
                count[4] += 1

        elif str(level) == "F":
            out_sheet["U" + str(count[5])] = key
            out_sheet["V" + str(count[5])] = value[1]
            out_sheet["W" + str(count[5])] = value[0]
            count[5] += 1

            if int(find_level[1]) == 18:
                count[5] += 1

    out_book.save(OUTPUT_FILENAME)
    os.remove(ORIGINAL_INPUT)


if __name__ == "__main__":
    main()