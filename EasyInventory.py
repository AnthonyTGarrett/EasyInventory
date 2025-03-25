#!/usr/bin/python3
# Author: Anthony Garrett
#
# Small script that will read in inventory location data from one spreadsheet
# and transfer that data to another spreadsheet
#
import os
import random as my_random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

#
# Input file name is set to the default file name used by the system that
#   generates the initial spreadsheet to be parsed
#

ORIGINAL_INPUT = "EXPORT.XLSX"
DATA_INPUT = "products.xlsx"

#
# Helper function to generate the dictionary that will hold the inventory count
# for the number of cases and pallets in the aisle in the base case with no
# special cases
#
# The value array is used as [pallet count, case count] and set to 0 for all
# as the default
#
# Input parameters are the aisle number as an int and the ending location for
# the aisle
#


def generate_inner_dict(aisle_number, ending_number, aisle_side):
    """
    Returns the default generated dictionary with the given aisle number
    and ending aisle location number.
    """

    locations_data = {}

    for i in range(1, ending_number):
        locations_data[str(aisle_number).zfill(2) + "-" + str((aisle_side + i)) + "-A"] = [
            0,
            0,
        ]
        locations_data[str(aisle_number).zfill(2) + "-" + str((aisle_side + i)) + "-B"] = [
            0,
            0,
        ]
        locations_data[str(aisle_number).zfill(2) + "-" + str((aisle_side + i)) + "-C"] = [
            0,
            0,
        ]
        locations_data[str(aisle_number).zfill(2) + "-" + str((aisle_side + i)) + "-D"] = [
            0,
            0,
        ]

    return locations_data


#
# Function to generate the dictionary that will hold the inventory count
# for the number of cases and pallets in the aisle.
#
# The value array is used as [pallet count, case count] and set to 0 for all
# as the default
#
# Special cases of uneven aisles and aisles that lack A levels on one side
# are dealt with in this function and base case aisles are handled in
# generate_inner_dict()
#
# Input paramenters are the aisle number as an int
#
def create_dictionary(aisle_number):
    """
    Returns the default generated dictionary with the given aisle number including special cases
    """
    locations = {}

    #
    # Decides what to do with each aisle number based on how the aisle is organized
    # in the warehouse. Each dictionary is created with the correct amount of valid
    # aisle locations all set to [0, 0]
    #
    #
    #
    # Aisle 1 only has 56 locations on each side
    #
    if aisle_number == 1:
        locations.update(generate_inner_dict(aisle_number, 57, 100))
        locations.update(generate_inner_dict(aisle_number, 57, 200))

    #
    # Aisle 2 only has 56 locations on 100 side
    # and 54 locations on 200 side
    #
    elif aisle_number == 2:
        locations.update(generate_inner_dict(aisle_number, 57, 100))
        locations.update(generate_inner_dict(aisle_number, 55, 200))

    #
    # Aisle 3 only has 54 locations on 100 side
    # and 56 locations on 200 side
    #
    elif aisle_number == 3:

        locations.update(generate_inner_dict(aisle_number, 55, 100))
        locations.update(generate_inner_dict(aisle_number, 57, 200))

    #
    # Aisle 4 has only 56 locations on the 100 side and 62 locations on the
    # 200 side
    #
    elif aisle_number == 4:
        locations.update(generate_inner_dict(aisle_number, 57, 100))
        locations.update(generate_inner_dict(aisle_number, 63, 200))

    #
    # Aisles 5 - 12 have base case locations with 62 on each side
    #
    elif 4 < aisle_number < 13:
        locations.update(generate_inner_dict(aisle_number, 63, 100))
        locations.update(generate_inner_dict(aisle_number, 63, 200))

    #
    # Aisle 13 is missing the A levels on the 200 side
    #
    elif aisle_number == 13:
        locations.update(generate_inner_dict(aisle_number, 63, 100))

        for j in range(1, 63):
            locations[str(aisle_number) + "-" + str((200 + j)) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" + str((200 + j)) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" + str((200 + j)) + "-D"] = [0, 0]

    #
    # Aisle 14 is missing the A levels on the 100 side
    #
    elif aisle_number == 14:
        for i in range(1, 63):
            locations[str(aisle_number) + "-" + str((100 + i)) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" + str((100 + i)) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" + str((100 + i)) + "-D"] = [0, 0]

        locations.update(generate_inner_dict(aisle_number, 63, 200))

    #
    # Aisles 15 - 26 have base case locations with 62 on each side
    #
    elif aisle_number <= 26:
        locations.update(generate_inner_dict(aisle_number, 63, 100))
        locations.update(generate_inner_dict(aisle_number, 63, 200))

    # Aisle 27 has six levels in parts of it so I need to create the dict
    # uniquely for this aisle
    elif aisle_number == 27:
        for i in range(1, 19):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-E"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-F"] = [0, 0]
        for i in range(19, 119):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]
        for i in range(119, 127):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-E"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-F"] = [0, 0]
        for i in range(127, 153):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]

    elif aisle_number == 28:
        for i in range(1, 69):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]

        for i in range(69, 73):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-E"] = [0, 0]

        for i in range(73, 145):
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-A"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-B"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-C"] = [0, 0]
            locations[str(aisle_number) + "-" +
                      str(i).zfill(3) + "-D"] = [0, 0]

    return locations


def main():

    # Loading the full pallet quantities from the product spreadsheet
    data_wb = load_workbook(DATA_INPUT)
    data_sheet = data_wb['Sheet1']

    data_products = data_sheet["A"]
    data_full_quantities = data_sheet["B"]
    data_products_dict = {}

    for i in range(len(data_products)):
        data_products_dict[data_products[i].value] = int(
            data_full_quantities[i].value)

    # Loading Full spreadsheet of aisle counts
    wb = load_workbook(ORIGINAL_INPUT)
    sheet = wb.active


    # Deleting the first row which contains headers for the columns
    sheet.delete_rows(1, 1)

    # Selecting the columns that hold the needed information about each location
    location_codes = sheet["A"]
    products = sheet["B"]
    batches = sheet["C"]
    actual_quantity = sheet["D"]
    handling_unit = sheet["G"]
    
    # Creating a list of the indexes of the number of location codes to randomly select 20 of the location codes.
    # This will create a sample of 20 pallets from the aisle to counting
    random_items = list(range(0, len(location_codes), 1))
    random_choices = my_random.sample(random_items, k=20)
    random_list = []
    
    for i in random_choices:
        random_list.append(
            [location_codes[i].value, products[i].value, batches[i].value, handling_unit[i].value, actual_quantity[i].value]
        )
    

    partials_list = []

    # Loop through actual_quantity list to compare actual quantity to total pallet quantity
    # of the specified product from data_products_dict
    for i in range(len(actual_quantity)):
        if actual_quantity[i].value < data_products_dict[int(products[i].value)]:
            partials_list.append(
                [location_codes[i].value, products[i].value, batches[i].value, handling_unit[i].value, actual_quantity[i].value])

    # Parsing out the aisle number from the location code
    code_parts = location_codes[0].value.split("-")[0]

    data_core = create_dictionary(int(code_parts))
    OUTPUT_FILENAME = "Aisle-" + code_parts + "-totals" + ".xlsx"

    # Cleaning excess spaces that are at the beginning of the location codes
    # in the cells
    for code in location_codes:
        code.value = code.value.strip()

    for count, code in enumerate(location_codes):
        data_core[code.value][0] += 1

        # Special case for pallets that are totes holding up to 3000 hot pockets
        # These are systematically entered as having 96 cases
        # No pallet has more than 220 cases except for these special tote
        # pallets
        if actual_quantity[count].value > 220:
            data_core[code.value][1] += 96
        else:
            data_core[code.value][1] += actual_quantity[count].value

    out_book = Workbook()

    out_sheet = out_book.active

    count = [1, 1, 1, 1, 1, 1]

    for key, value in data_core.items():
        find_level = key.split("-")
        level = find_level[2]

        if str(level) == "A":
            out_sheet["A" + str(count[0])] = key
            out_sheet["B" + str(count[0])] = value[1]
            out_sheet["C" + str(count[0])] = value[0]
            count[0] += 1

        elif str(level) == "B":
            out_sheet["E" + str(count[1])] = key
            out_sheet["F" + str(count[1])] = value[1]
            out_sheet["G" + str(count[1])] = value[0]
            count[1] += 1

        elif str(level) == "C":
            out_sheet["I" + str(count[2])] = key
            out_sheet["J" + str(count[2])] = value[1]
            out_sheet["K" + str(count[2])] = value[0]
            count[2] += 1

        elif str(level) == "D":
            out_sheet["M" + str(count[3])] = key
            out_sheet["N" + str(count[3])] = value[1]
            out_sheet["O" + str(count[3])] = value[0]
            count[3] += 1

        elif str(level) == "E":
            out_sheet["Q" + str(count[4])] = key
            out_sheet["R" + str(count[4])] = value[1]
            out_sheet["S" + str(count[4])] = value[0]
            count[4] += 1

            if int(find_level[1]) == 18:
                count[4] += 1

        elif str(level) == "F":
            out_sheet["U" + str(count[5])] = key
            out_sheet["V" + str(count[5])] = value[1]
            out_sheet["W" + str(count[5])] = value[0]
            count[5] += 1

            if int(find_level[1]) == 18:
                count[5] += 1

        partial_out_book = Workbook()

    partial_out_sheet = partial_out_book.active

    aisle_number = location_codes[0].value.split("-")[0]
    PARTIAL_OUTPUT_FILENAME = "Aisle-" + aisle_number + "-partials" + ".xlsx"

    # Setting Up headers for the spreadsheet
    partial_out_sheet["A1"] = "Storage Bin"
    partial_out_sheet.column_dimensions['A'].width = 12
    partial_out_sheet.row_dimensions[1].height = 25

    partial_out_sheet["B1"] = "Product"
    partial_out_sheet.column_dimensions['B'].width = 12
    
    partial_out_sheet["C1"] = "Batch"
    partial_out_sheet.column_dimensions['C'].width = 15

    partial_out_sheet["D1"] = "Handling Unit"
    partial_out_sheet.column_dimensions['D'].width = 20

    partial_out_sheet["E1"] = "Quantity"
    partial_out_sheet.column_dimensions['E'].width = 9

    for count, partial in enumerate(partials_list, start=2):
        partial_out_sheet["A" + str(count)] = partial[0]
        partial_out_sheet["B" + str(count)] = partial[1]
        partial_out_sheet["C" + str(count)] = partial[2]
        partial_out_sheet["D" + str(count)] = partial[3]
        partial_out_sheet["E" + str(count)] = partial[4]

    for cell in partial_out_sheet['A:A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['B:B']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['C:C']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['D:D']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in partial_out_sheet['E:E']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
# 
# 
# Random Pallet sample output file setup
# 
# 

    random_out_book = Workbook()

    random_out_sheet = random_out_book.active

    aisle_number = location_codes[0].value.split("-")[0]
    RANDOM_OUTPUT_FILENAME = "Aisle-" + aisle_number + "-samples" + ".xlsx"

    # Setting Up headers for the spreadsheet
    random_out_sheet["A1"] = "Storage Bin"
    random_out_sheet.column_dimensions['A'].width = 12
    random_out_sheet.row_dimensions[1].height = 25

    random_out_sheet["B1"] = "Product"
    random_out_sheet.column_dimensions['B'].width = 12
    
    random_out_sheet["C1"] = "Batch"
    random_out_sheet.column_dimensions['C'].width = 15

    random_out_sheet["D1"] = "Handling Unit"
    random_out_sheet.column_dimensions['D'].width = 20

    random_out_sheet["E1"] = "Quantity"
    random_out_sheet.column_dimensions['E'].width = 9

    for count, random in enumerate(random_list, start=2):
        random_out_sheet["A" + str(count)] = random[0]
        random_out_sheet["B" + str(count)] = random[1]
        random_out_sheet["C" + str(count)] = random[2]
        random_out_sheet["D" + str(count)] = random[3]
        random_out_sheet["E" + str(count)] = random[4]

    for cell in random_out_sheet['A:A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in random_out_sheet['B:B']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in random_out_sheet['C:C']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in random_out_sheet['D:D']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for cell in random_out_sheet['E:E']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        
    
    random_out_book.save(RANDOM_OUTPUT_FILENAME)
    partial_out_book.save(PARTIAL_OUTPUT_FILENAME)
    out_book.save(OUTPUT_FILENAME)
    os.remove(ORIGINAL_INPUT)


if __name__ == "__main__":
    main()
